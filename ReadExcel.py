#encoding=latin1

# Tool to read Excel data and create issues to Jira
#
# Author mika.nokka1@gmail.com 4.2.2022 
#
# python3
#
#from __future__ import unicode_literals

import openpyxl 
import sys, logging
import argparse
#import re
from collections import defaultdict
#from author import Authenticate  # no need to use as external command
#from author import DoJIRAStuff
import glob
import json # for json dumo
#from sqlalchemy.sql.expression import false
import re
import time
import os
import datetime
from IssueCreator import CreateIssue
from Authorization import Authenticate
from Authorization import DoJIRAStuff
import datetime


######################################################################################
#CONFIGURATIONS
__version__ = "0.5"

start = datetime.datetime.now()

#
# FOR CODE EMBEDDED  SETTINGS ====>  SEE Parse FUNCTION
#
logging.basicConfig(level=logging.DEBUG) # INFO IF calling from Groovy, this must be set logging level DEBUG in Groovy side order these to be written out

#######################################################################################

start = datetime.datetime.now()


def main(argv):
    
    JIRASERVICE=""
    JIRAPROJECT=""
    PSWD=''
    USER=''
  
    logging.debug ("--Python starting Excel reading --") 

 
    parser = argparse.ArgumentParser(usage="""
    {1}    Version:{0}     -  mika.nokka1@gmail.com
    
    USAGE:
    -filepath  | -p <Path to Excel file directory>
    -filename   | -n <Excel filename>

    """.format(__version__,sys.argv[0]))

    parser.add_argument('-f','--filepath', help='<Path to Excel file directory>')
    parser.add_argument('-n','--filename', help='<Excel filename>')
   # parser.add_argument('-m','--subfilename', help='<Subtasks Excel filename>')
    parser.add_argument('-v','--version', help='<Version>', action='store_true')
    
    parser.add_argument('-w','--password', help='<JIRA password or token>')
    parser.add_argument('-u','--user', help='<JIRA user or email address if Jira cloud token used>')
    parser.add_argument('-s','--service', help='<JIRA service>')
    parser.add_argument('-p','--project', help='<JIRA project key>')
   
    #parser.add_argument('-a','--attachemnts', help='<Attachment directory>')
        
    args = parser.parse_args()
    
    if args.version:
        print ("Tool version: {0}".format(__version__))
        sys.exit(2)    
           
    filepath = args.filepath or ''
    filename = args.filename or ''
    #subfilename=args.subfilename or ''
    
    JIRASERVICE = args.service or ''
    JIRAPROJECT = args.project or ''
    PSWD= args.password or ''
    USER= args.user or ''
    #ATTACHDIR=args.attachemnts or ''
    
    # quick old-school way to check needed parameters
    if (filepath=='' or  filename=='' or JIRASERVICE=='' or  JIRAPROJECT==''  or PSWD=='' or USER=='' ):
        parser.print_help()
        sys.exit(2)
        


    Parse(filepath, filename,JIRASERVICE,JIRAPROJECT,PSWD,USER)


############################################################################################################################################
# Parse excel and create dictionary of
# Jiraissue data
#  
# Uses hardcoded sheet/column value,change accordingly

def Parse(filepath, filename,JIRASERVICE,JIRAPROJECT,PSWD,USER):
    

    ####################################################################################
    # CONFIGURATIONS ####
    PROD=True # False / True   #false skips issue creation and other jira operations
    ENV="PROD" # or "PROD" or "DEV", sets the custom field IDs 
    AUTH=True # True / False ,for  jira authorizations
    DRY="on" # on==dry run, dont do   off=do everything THIS IS THE ONE FLAG TO RULE THEM ALL
    # END OF CONFIGURATIONS ############################################################
    
    # flag to indicate whether issue under operations have been already created to Jira
    IMPORT=False
    
    logging.info ("Filepath: %s     Filename:%s" %(filepath ,filename))
    files=filepath+"/"+filename
    logging.info ("Excel file:{0}".format(files))
   
    Issues=defaultdict(dict) 
    MainSheet="Sheet1" 

    try:
        wb= openpyxl.load_workbook(files)
    except: 
      logging.error ("Can't open excel. EXITING")
      sys.exit(5)
       
    types=type(wb)
    logging.debug ("Type:{0}".format(types))
    sheets=wb.sheetnames
    logging.debug ("Sheets:{0}".format(sheets))
   
    CurrentSheet=wb[MainSheet] 
    logging.debug ("CurrentSheet:{0}".format(CurrentSheet))
    #logging.debug ("Column A, row:{0}".format(CurrentSheet['A4'].value))




    ##############################################################################
    #CONFIGURATIONS AND EXCEL COLUMN MAPPINGS
    #
    DATASTARTSROW=4 # data section starting line in the excel sheet
    A=1 
    B=2 
    C=3 
    D=4 
    E=5
    F=6 
    G=7
    H=8
    I=9
    J=10
    K=11
    L=12
    M=13
    N=14
    O=15
    
    
    
    ####################################################################################################################
    # Go through main excel sheet for main issue keys (and contents findings)
    # Create dictionary structure
    # NOTE: Uses hardcoded sheet/column values. A is assumed to hold KEY, used to create dictionary entry
    # NOTE: As this handles first sheet, using used row/cell reading (buggy, works only for first sheet) 
    #
    i=DATASTARTSROW # brute force row indexing
    ENDROW=(CurrentSheet.max_row) # to prevent off-by-one in the end of sheet, also excel needs deleting of empty end line1
    logging.debug ("STARTROW:{0} ENDROW:{1}".format(DATASTARTSROW,ENDROW))

    for row in CurrentSheet[('A{}:A{}'.format(DATASTARTSROW,ENDROW))]:  # go trough all column A (KEY) rows
        for mycell in row:
            KEY=mycell.value # column A
            logging.debug  ("KEY:{0} data in ROW:{1}".format(KEY,i))
            Issues[KEY]={} # add KEY to dictionary as master key 
            
            #Hardcoded ovalue picking and dictionary settings operations
        
            Reference=(CurrentSheet.cell(row=i, column=A).value)
            if not Reference:
                Reference="not defined"
            Issues[KEY]["Reference"] = Reference
        
        
            JIRASUMMARY=(CurrentSheet.cell(row=i, column=B).value)
            if not JIRASUMMARY:
                JIRASUMMARY="not defined"
            Issues[KEY]["JIRASUMMARY"] = JIRASUMMARY
            
            # selection list, several choices
            Domain=(CurrentSheet.cell(row=i, column=C).value)
            if not Domain:
                Domain="not defined"
            Issues[KEY]["Domain"] = Domain
            
            # selection list "Speculative" + "xxx" + "zzz"
            ExploitStatus=(CurrentSheet.cell(row=i, column=D).value)
            if not ExploitStatus:
                Authorization="not defined"
            Issues[KEY]["ExploitStatus"] = ExploitStatus
            
            #0-10 float
            Base=(CurrentSheet.cell(row=i, column=E).value)
            if not Base:
                Base=""
            Issues[KEY]["Base"] = Base
           
            #0-10 float
            Impact=(CurrentSheet.cell(row=i, column=F).value)
            if not Impact:
                Impact=""
            Issues[KEY]["Impact"] = Impact
          
            #0-10 float
            Exploitability=(CurrentSheet.cell(row=i, column=G).value)
            if not Exploitability:
                Impact=""
            Issues[KEY]["Exploitability"] = Exploitability
            
            #string, security attack vector
            Vector=(CurrentSheet.cell(row=i, column=H).value)
            if not Impact:
                Impact=""
            Issues[KEY]["Vector"] = Vector
          
            # ? Yes No
            Authencity=(CurrentSheet.cell(row=i, column=I).value)
            if not Impact:
                Impact="?"
            Issues[KEY]["Authencity"] = Authencity
            
            # ? Yes No
            Integrity=(CurrentSheet.cell(row=i, column=J).value)
            if not Integrity:
                Integrity="?"
            Issues[KEY]["Integrity"] = Integrity
            
            # ? Yes No
            Nonrepudiability=(CurrentSheet.cell(row=i, column=K).value)
            if not Nonrepudiability:
                Nonrepudiability="?"
            Issues[KEY]["Nonrepudiability"] = Nonrepudiability
            
              # ? Yes No
            Confidentiality=(CurrentSheet.cell(row=i, column=L).value)
            if not Confidentiality:
                Confidentiality="?"
            Issues[KEY]["Confidentiality"] = Confidentiality
            
              # ? Yes No
            Availability=(CurrentSheet.cell(row=i, column=M).value)
            if not Availability:
                Availability="?"
            Issues[KEY]["Availability"] = Availability
            
            # ? Yes No
            Authorization=(CurrentSheet.cell(row=i, column=N).value)
            if not Authorization:
                Authorization="?"
            Issues[KEY]["Authorization"] = Authorization
            
            DESCRIPTION=(CurrentSheet.cell(row=i, column=O).value)
            if not DESCRIPTION:
                DESCRIPTION="?"
            Issues[KEY]["DESCRIPTION"] = DESCRIPTION
            
            
            
            logging.debug("---------------------------------------------------")
            i=i+1
            
          
    logging.debug (Issues)
    logging.debug (Issues.items()) 
    #logging.debug((json.dumps(Issues, indent=4, sort_keys=True)))
    logging.debug((json.dumps(Issues, indent=4 )))
    
    #key=18503 # check if this key exists
    #if key in Issues:
    #    print "EXISTS"
    #else:
    #    print "NOT THERE"
    #for key, value in Issues.iteritems() :
    #    print key, value

    

    ##########################################################################################################################
    # Create main issues
    
    if (AUTH==True):    
        Authenticate(JIRASERVICE,PSWD,USER)
        jira=DoJIRAStuff(USER,PSWD,JIRASERVICE)
    else:
        print ("Simulated execution only")

    #create Jira issues
    i=1
    for key, value in Issues.items() :
        #if (i>20):
        #    print "EXIT DUE i"
        #    sys.exit(1)
        KEYVALUE=(key,value)
        KEY=key
        print ("ORIGINAL ISSUE KEY:{0}\nVALUE:{1}".format(KEY,KEYVALUE))
        
        # TODO: check if issue has been imported in previous import attempts
        #if (ENV=="DEV"):
        #    JQLQuery="project = {0}  and cf[12900]  ~ {1}".format(JIRAPROJECT,key) # check key in Jira
        #    results=jira.search_issues(JQLQuery, maxResults=3000)
        #elif (ENV=="PROD"):
            #print "NOT IMPLEMENTED PROD CODE"
            #sys.exit(1)
        #    JQLQuery="project = {0}  and cf[12900]  ~ {1}".format(JIRAPROJECT,key) # check key in Jira
        #    results=jira.search_issues(JQLQuery, maxResults=3000)   
        #else:
        #    print ("ENV SET WRONG")
        #    sys.exit(1)
               
        #if (len(results) > 0):
        #    print ("Key:{0} exists in Jira already".format(key))
        #    print ("Query result:{0}".format(results))
        #    IMPORT=False
        #else:
        #    print ("Key:{0} is a NEW Key,going to import".format(key))
        #    print ("Query result:{0}".format(results))    
        #    IMPORT=True
             
            
        ####### left old ones as examples of data formatting ######################
        # ISSUETYPENW=str((Issues[key]["ISSUE_TYPENW"]).encode('utf-8'))  # str casting needed      
        # JIRASUMMARY=JIRASUMMARY.replace("\n", " ") # Perl used to have chomp, this was only Python way to do this

        #DECKNW=(Issues[key]["DECKNW"])
        #if (DECKNW is None):
        #    DECKNW=Issues[key]["DECKNW"] #to keep None object??
        #    DECKNW="1" # just set some random default value
        #else: 
        #    DECKNW=str((Issues[key]["DECKNW"]))  # str casting needed
        #print ("DECKNW:{0}".format(DECKNW)) 
               
        #set custom field setting variables
        
        
 
               
               
        #temp settings to test 
        ISSUETYPE="Epic"
        PRIORITY="Low"
        
        #set values from excel based dictionary
        JIRASUMMARY=Issues[key]["JIRASUMMARY"]
        Reference=Issues[KEY]["Reference"] 
        Domain = Issues[KEY]["Domain"]  
        ExploitStatus = Issues[KEY]["ExploitStatus"] 
        Base = Issues[KEY]["Base"]  
        Impact = Issues[KEY]["Impact"] 
        
        Exploitability =  Issues[KEY]["Exploitability"] = Exploitability
        Vector = Issues[KEY]["Vector"] 
        Authencity= Issues[KEY]["Authencity"] 
        Integrity=  Issues[KEY]["Integrity"] 
        Nonrepudiability= Issues[KEY]["Nonrepudiability"] 
        Confidentiality= Issues[KEY]["Confidentiality"] 
        Availability=  Issues[KEY]["Availability"] 
        Authorization= Issues[KEY]["Authorization"] 

        DESCRIPTION=Issues[key]["DESCRIPTION"]
                                                                         
        

        
        
        
        
        JIRASUMMARY=JIRASUMMARY[:254] ## summary max length is 255
        
        IMPORT=True # temp setting

        if (PROD==True):
            if (IMPORT==True):
                if (DRY=="off"):
                   IssueID=CreateIssue(jira,JIRAPROJECT,JIRASUMMARY,ISSUETYPE,PRIORITY,DESCRIPTION,Reference,Domain,ExploitStatus,Base,Impact,Exploitability,Vector,Authencity,Integrity,Nonrepudiability,Confidentiality,Availability,Authorization  )
                   print ("Created issue:{0}  OK".format(IssueID))
                   print ("---------------PAUSED 0.2 secs --------------------------------------------")
                   time.sleep(0.2) 
                

                
                elif (DRY=="on"):       
                    print (" **** Dryrun mode: I would have Created issue  *****")
                    #print ("JIRAPROJECT:{0},JIRASUMMARY:{1},ISSUETYPE:{2},PRIORITY:{3},DESCRIPTION:{4})".format(JIRAPROJECT,JIRASUMMARY,ISSUETYPE,PRIORITY,DESCRIPTION)) 
                    print ("-----------------------------------------------------------")


            else:
                print ("Issue exists in Jira. Did nothing")
        else:
           print ("--> SKIPPED ISSUE CREATION") 
      

        
        i=i+1    
    
      
    end = datetime.datetime.now()
    totaltime=end-start
    seconds=totaltime.total_seconds()
    print ("---> Time taken:{0} seconds".format(totaltime))
        
#############################################################################


    
logging.debug ("--Python exiting--")
if __name__ == "__main__":
    main(sys.argv[1:]) 